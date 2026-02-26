"""
FastAPI アプリケーション

Excel帳票エンジンのメインアプリケーション定義です。
"""

import logging
import os
import uuid
from contextlib import asynccontextmanager
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any, Optional

from fastapi import FastAPI, Query
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from fpdf import FPDF
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from pydantic import BaseModel

from app.api.router import api_router
from app.config import settings
from app.database import dispose_db, init_db

# ロギング設定
logging.basicConfig(
    level=settings.LOG_LEVEL,
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
)

logger = logging.getLogger(__name__)


@asynccontextmanager
async def lifespan(app: FastAPI):
    """
    アプリケーションのライフサイクル管理

    起動時と終了時の処理を定義します。
    """
    # 起動処理
    logger.info(f"アプリケーション起動: {settings.APP_NAME} v{settings.APP_VERSION}")
    logger.info(f"DEBUG={settings.DEBUG}, LOG_LEVEL={settings.LOG_LEVEL}")
    logger.info(f"STORAGE_TYPE={settings.STORAGE_TYPE}")

    try:
        await init_db()
        logger.info("データベース初期化完了")
    except Exception as e:
        logger.error(f"データベース初期化エラー: {str(e)}")
        raise

    yield

    # 終了処理
    logger.info("アプリケーション終了中...")
    try:
        await dispose_db()
        logger.info("データベース接続をクローズしました")
    except Exception as e:
        logger.error(f"データベース接続クローズエラー: {str(e)}")

    logger.info("アプリケーション終了完了")


# FastAPI アプリケーション作成
app = FastAPI(
    title=settings.APP_NAME,
    description="Excel帳票エンジン - テンプレートベースのExcelファイル生成システム",
    version=settings.APP_VERSION,
    openapi_url=settings.API_OPENAPI_URL,
    docs_url=settings.API_DOCS_URL,
    redoc_url=settings.API_REDOC_URL,
    lifespan=lifespan,
)

# CORS ミドルウェア設定
app.add_middleware(
    CORSMiddleware,
    allow_origins=settings.ALLOWED_ORIGINS,
    allow_credentials=settings.ALLOWED_CREDENTIALS,
    allow_methods=settings.ALLOWED_METHODS,
    allow_headers=settings.ALLOWED_HEADERS,
)


# ============================================================
# ルートエンドポイント（Railway ステータス確認用）
# ============================================================

@app.get("/", tags=["status"])
async def root():
    """Railway ステータス確認"""
    return {"status": "Postgres connected"}


# ============================================================
# ヘルスチェック
# ============================================================

@app.get("/health", tags=["health"])
async def health_check():
    """
    ヘルスチェック

    アプリケーションが正常に動作しているか確認します。
    """
    return {
        "status": "healthy",
        "app": settings.APP_NAME,
        "version": settings.APP_VERSION,
    }


# ============================================================
# テンプレートベース Excel 生成 ヘルパー
# ============================================================

# テンプレートファイルパス（Railway デプロイ時は COPY 済み）
ESTIMATE_TEMPLATE_PATH = Path(__file__).resolve().parent.parent / "storage" / "templates" / "estimate_v2.xlsx"


def _find_merge_top_left(ws: Worksheet, cell_ref: str) -> Optional[str]:
    """マージセルの左上セルを検出"""
    for merged_range in ws.merged_cells.ranges:
        if cell_ref in merged_range:
            return merged_range.start_cell.coordinate
    return None


def _safe_write(ws: Worksheet, cell_ref: str, value: Any) -> None:
    """マージセル対応の安全書き込み"""
    top_left = _find_merge_top_left(ws, cell_ref)
    ws[top_left or cell_ref] = value


# ============================================================
# 見積書 Excel 生成 API（テンプレートベース）
# ============================================================

# セルマッピング定義（init_db.py と同じ36マッピング）
ESTIMATE_CELL_MAPPINGS = {
    "customer_name":     "B5",
    "estimate_date":     "N5",
    "total_with_tax_2":  "K10",
    "body_price":        "N20",
    "option_total":      "N22",
    "futai_total":       "N24",
    "contract_amount":   "N27",
    "tax":               "N30",
    "total_with_tax":    "N33",
    "body_price_detail": "X65",
    "opt_01_amount":     "U69",
    "opt_02_amount":     "U70",
    "opt_03_amount":     "U71",
    "opt_04_amount":     "U72",
    "opt_05_amount":     "U73",
    "opt_06_amount":     "U74",
    "opt_07_amount":     "U75",
    "opt_08_amount":     "U76",
    "opt_09_amount":     "U77",
    "opt_10_amount":     "U78",
    "option_subtotal":   "N82",
    "futai_01_amount":   "U86",
    "futai_02_amount":   "U87",
    "futai_03_amount":   "U88",
    "futai_04_amount":   "U89",
    "futai_05_amount":   "U90",
    "futai_06_amount":   "U91",
    "futai_07_amount":   "U92",
    "futai_08_amount":   "U93",
    "futai_subtotal":    "N96",
    "floor_1f_sqm":      "E48",
    "floor_1f_tsubo":    "I48",
    "floor_2f_sqm":      "E49",
    "floor_2f_tsubo":    "I49",
    "floor_total_sqm":   "E51",
    "floor_total_tsubo": "I51",
}


class EstimateRequest(BaseModel):
    """見積書Excel生成リクエスト（テンプレートベース）"""
    # 基本情報
    customer_name: str = ""
    estimate_date: str = ""
    # 金額
    body_price: Optional[float] = None
    option_total: Optional[float] = None
    futai_total: Optional[float] = None
    contract_amount: Optional[float] = None
    tax: Optional[float] = None
    total_with_tax: Optional[float] = None
    total_with_tax_2: Optional[float] = None
    body_price_detail: Optional[float] = None
    # オプション工事 (10行)
    opt_01_amount: Optional[float] = None
    opt_02_amount: Optional[float] = None
    opt_03_amount: Optional[float] = None
    opt_04_amount: Optional[float] = None
    opt_05_amount: Optional[float] = None
    opt_06_amount: Optional[float] = None
    opt_07_amount: Optional[float] = None
    opt_08_amount: Optional[float] = None
    opt_09_amount: Optional[float] = None
    opt_10_amount: Optional[float] = None
    option_subtotal: Optional[float] = None
    # 付帯工事 (8行)
    futai_01_amount: Optional[float] = None
    futai_02_amount: Optional[float] = None
    futai_03_amount: Optional[float] = None
    futai_04_amount: Optional[float] = None
    futai_05_amount: Optional[float] = None
    futai_06_amount: Optional[float] = None
    futai_07_amount: Optional[float] = None
    futai_08_amount: Optional[float] = None
    futai_subtotal: Optional[float] = None
    # 床面積
    floor_1f_sqm: Optional[float] = None
    floor_1f_tsubo: Optional[float] = None
    floor_2f_sqm: Optional[float] = None
    floor_2f_tsubo: Optional[float] = None
    floor_total_sqm: Optional[float] = None
    floor_total_tsubo: Optional[float] = None


class GenerateRequest(BaseModel):
    """見積書Excel生成リクエスト（基本3フィールド互換）"""
    project_name: str = ""
    version: str = ""
    author: str = ""


class SpecSheetRequest(BaseModel):
    """仕様書Excel生成リクエスト（全フィールド対応）"""
    project: str = ""
    category: str = ""
    structure: str = ""
    floors: str = ""
    foundation: str = ""
    roofing: str = ""
    exterior: str = ""
    insulation: str = ""
    window: str = ""
    flooring: str = ""
    kitchen: str = ""
    bath: str = ""
    toilet: str = ""
    aircon: str = ""
    ventilation: str = ""
    note: str = ""
    date: str = ""
    status: str = ""


# 共通スタイル定義
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

_TITLE_FONT = Font(bold=True, size=16, color="0D9488")
_SECTION_FONT = Font(bold=True, size=12, color="FFFFFF")
_HEADER_FONT = Font(bold=True, size=10)
_VALUE_FONT = Font(size=10)
_HEADER_FILL = PatternFill(start_color="E0F2F1", end_color="E0F2F1", fill_type="solid")
_SECTION_FILL = PatternFill(start_color="0D9488", end_color="0D9488", fill_type="solid")
_NOTE_FILL = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
_THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)
_CENTER = Alignment(horizontal="center", vertical="center")
_LEFT = Alignment(vertical="center")


def _build_spec_workbook(data: dict) -> Workbook:
    """仕様書Excelワークブックを構築"""
    wb = Workbook()
    ws = wb.active
    ws.title = "仕様書"

    # 列幅
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 45

    # 印刷設定
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = "portrait"
    ws.print_options.horizontalCentered = True
    ws.page_margins.left = 0.6
    ws.page_margins.right = 0.6

    row = 1

    # === タイトル ===
    ws.merge_cells("A1:B1")
    title_cell = ws.cell(row=1, column=1, value="建 築 仕 様 書")
    title_cell.font = _TITLE_FONT
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40
    row = 3

    # === 基本情報セクション ===
    basic_fields = [
        ("工事名", data.get("project", "")),
        ("仕様区分", data.get("category", "")),
        ("作成日", data.get("date", datetime.now().strftime("%Y-%m-%d"))),
        ("ステータス", data.get("status", "")),
    ]
    ws.merge_cells(f"A{row}:B{row}")
    sec = ws.cell(row=row, column=1, value="基本情報")
    sec.font = _SECTION_FONT
    sec.fill = _SECTION_FILL
    sec.alignment = _CENTER
    ws.cell(row=row, column=2).fill = _SECTION_FILL
    ws.row_dimensions[row].height = 28
    row += 1

    for label, val in basic_fields:
        ws.cell(row=row, column=1, value=label).font = _HEADER_FONT
        ws.cell(row=row, column=1).fill = _HEADER_FILL
        ws.cell(row=row, column=1).border = _THIN_BORDER
        ws.cell(row=row, column=1).alignment = _LEFT
        ws.cell(row=row, column=2, value=val).font = _VALUE_FONT
        ws.cell(row=row, column=2).border = _THIN_BORDER
        ws.cell(row=row, column=2).alignment = _LEFT
        ws.row_dimensions[row].height = 24
        row += 1

    row += 1

    # === 構造・躯体セクション ===
    struct_fields = [
        ("構造種別", data.get("structure", "")),
        ("階数", data.get("floors", "")),
        ("基礎形式", data.get("foundation", "")),
        ("屋根材", data.get("roofing", "")),
        ("外壁材", data.get("exterior", "")),
    ]
    ws.merge_cells(f"A{row}:B{row}")
    sec = ws.cell(row=row, column=1, value="構造・躯体")
    sec.font = _SECTION_FONT
    sec.fill = _SECTION_FILL
    sec.alignment = _CENTER
    ws.cell(row=row, column=2).fill = _SECTION_FILL
    ws.row_dimensions[row].height = 28
    row += 1

    for label, val in struct_fields:
        ws.cell(row=row, column=1, value=label).font = _HEADER_FONT
        ws.cell(row=row, column=1).fill = _HEADER_FILL
        ws.cell(row=row, column=1).border = _THIN_BORDER
        ws.cell(row=row, column=1).alignment = _LEFT
        ws.cell(row=row, column=2, value=val).font = _VALUE_FONT
        ws.cell(row=row, column=2).border = _THIN_BORDER
        ws.cell(row=row, column=2).alignment = _LEFT
        ws.row_dimensions[row].height = 24
        row += 1

    row += 1

    # === 断熱・開口部セクション ===
    insul_fields = [
        ("断熱仕様", data.get("insulation", "")),
        ("サッシ・窓", data.get("window", "")),
    ]
    ws.merge_cells(f"A{row}:B{row}")
    sec = ws.cell(row=row, column=1, value="断熱・開口部")
    sec.font = _SECTION_FONT
    sec.fill = _SECTION_FILL
    sec.alignment = _CENTER
    ws.cell(row=row, column=2).fill = _SECTION_FILL
    ws.row_dimensions[row].height = 28
    row += 1

    for label, val in insul_fields:
        ws.cell(row=row, column=1, value=label).font = _HEADER_FONT
        ws.cell(row=row, column=1).fill = _HEADER_FILL
        ws.cell(row=row, column=1).border = _THIN_BORDER
        ws.cell(row=row, column=1).alignment = _LEFT
        ws.cell(row=row, column=2, value=val).font = _VALUE_FONT
        ws.cell(row=row, column=2).border = _THIN_BORDER
        ws.cell(row=row, column=2).alignment = _LEFT
        ws.row_dimensions[row].height = 24
        row += 1

    row += 1

    # === 内装・設備セクション ===
    equip_fields = [
        ("床材（LDK）", data.get("flooring", "")),
        ("キッチン", data.get("kitchen", "")),
        ("UB", data.get("bath", "")),
        ("トイレ", data.get("toilet", "")),
    ]
    ws.merge_cells(f"A{row}:B{row}")
    sec = ws.cell(row=row, column=1, value="内装・設備")
    sec.font = _SECTION_FONT
    sec.fill = _SECTION_FILL
    sec.alignment = _CENTER
    ws.cell(row=row, column=2).fill = _SECTION_FILL
    ws.row_dimensions[row].height = 28
    row += 1

    for label, val in equip_fields:
        ws.cell(row=row, column=1, value=label).font = _HEADER_FONT
        ws.cell(row=row, column=1).fill = _HEADER_FILL
        ws.cell(row=row, column=1).border = _THIN_BORDER
        ws.cell(row=row, column=1).alignment = _LEFT
        ws.cell(row=row, column=2, value=val).font = _VALUE_FONT
        ws.cell(row=row, column=2).border = _THIN_BORDER
        ws.cell(row=row, column=2).alignment = _LEFT
        ws.row_dimensions[row].height = 24
        row += 1

    row += 1

    # === 空調・換気セクション ===
    hvac_fields = [
        ("空調方式", data.get("aircon", "")),
        ("換気方式", data.get("ventilation", "")),
    ]
    ws.merge_cells(f"A{row}:B{row}")
    sec = ws.cell(row=row, column=1, value="空調・換気")
    sec.font = _SECTION_FONT
    sec.fill = _SECTION_FILL
    sec.alignment = _CENTER
    ws.cell(row=row, column=2).fill = _SECTION_FILL
    ws.row_dimensions[row].height = 28
    row += 1

    for label, val in hvac_fields:
        ws.cell(row=row, column=1, value=label).font = _HEADER_FONT
        ws.cell(row=row, column=1).fill = _HEADER_FILL
        ws.cell(row=row, column=1).border = _THIN_BORDER
        ws.cell(row=row, column=1).alignment = _LEFT
        ws.cell(row=row, column=2, value=val).font = _VALUE_FONT
        ws.cell(row=row, column=2).border = _THIN_BORDER
        ws.cell(row=row, column=2).alignment = _LEFT
        ws.row_dimensions[row].height = 24
        row += 1

    row += 1

    # === 特記事項 ===
    note = data.get("note", "")
    if note:
        ws.merge_cells(f"A{row}:B{row}")
        sec = ws.cell(row=row, column=1, value="特記事項")
        sec.font = _SECTION_FONT
        sec.fill = _SECTION_FILL
        sec.alignment = _CENTER
        ws.cell(row=row, column=2).fill = _SECTION_FILL
        ws.row_dimensions[row].height = 28
        row += 1
        ws.merge_cells(f"A{row}:B{row}")
        note_cell = ws.cell(row=row, column=1, value=note)
        note_cell.font = _VALUE_FONT
        note_cell.fill = _NOTE_FILL
        note_cell.border = _THIN_BORDER
        note_cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row].height = 60
        row += 1

    row += 2

    # === 承認欄 ===
    ws.merge_cells(f"A{row}:B{row}")
    ws.cell(row=row, column=1, value="承認").font = Font(bold=True, size=10, color="666666")
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="right")
    row += 1
    for label in ["設計", "施工", "承認"]:
        ws.cell(row=row, column=1, value=label).font = Font(size=9, color="999999")
        ws.cell(row=row, column=1).alignment = _CENTER
        ws.cell(row=row, column=1).border = _THIN_BORDER
        ws.cell(row=row, column=2).border = _THIN_BORDER
        ws.row_dimensions[row].height = 40
        row += 1

    return wb


@app.post("/generate", tags=["generate"])
async def generate_estimate_excel(req: GenerateRequest):
    """
    見積書Excelを生成してダウンロード（基本3フィールド互換版）

    アップロード済みテンプレート（estimate_v2.xlsx）を読み込み、
    データを埋め込んでFileResponseで返します。
    """
    if not ESTIMATE_TEMPLATE_PATH.exists():
        logger.error(f"テンプレートが見つかりません: {ESTIMATE_TEMPLATE_PATH}")
        return {"error": "テンプレートファイルが見つかりません"}

    wb = load_workbook(str(ESTIMATE_TEMPLATE_PATH), data_only=False)
    ws = wb.active

    # 基本3フィールドをセルに書き込み
    _safe_write(ws, "B5", req.project_name)  # お客様名/工事名
    _safe_write(ws, "N5", datetime.now().strftime("%Y-%m-%d"))  # 見積日

    filename = f"estimate_{uuid.uuid4().hex[:8]}.xlsx"
    filepath = os.path.join("/tmp", filename)
    wb.save(filepath)

    safe_name = req.project_name.replace("/", "_").replace("\\", "_") if req.project_name else "見積書"
    download_name = f"見積書_{safe_name}_{datetime.now().strftime('%Y%m%d')}.xlsx"

    return FileResponse(
        path=filepath,
        filename=download_name,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.post("/generate-estimate", tags=["generate"])
async def generate_full_estimate_excel(req: EstimateRequest):
    """
    見積書Excelを生成してダウンロード（全36フィールド対応）

    アップロード済みテンプレート（estimate_v2.xlsx）を読み込み、
    36個のセルマッピングに基づいてデータを埋め込みます。
    マージセル（180箇所）にも安全に対応。
    """
    if not ESTIMATE_TEMPLATE_PATH.exists():
        logger.error(f"テンプレートが見つかりません: {ESTIMATE_TEMPLATE_PATH}")
        return {"error": "テンプレートファイルが見つかりません"}

    wb = load_workbook(str(ESTIMATE_TEMPLATE_PATH), data_only=False)
    ws = wb.active

    # リクエストデータをdict化
    data = req.model_dump(exclude_none=True)

    # セルマッピングに基づいてデータを埋め込み
    for field_name, cell_ref in ESTIMATE_CELL_MAPPINGS.items():
        if field_name in data and data[field_name] is not None:
            value = data[field_name]
            # 日付フィールドの処理
            if field_name == "estimate_date" and isinstance(value, str) and value:
                try:
                    value = datetime.fromisoformat(value)
                except ValueError:
                    pass  # ISO形式でなければ文字列のまま
            _safe_write(ws, cell_ref, value)

    logger.info(f"見積書生成完了: {len(data)}フィールド埋め込み")

    filename = f"estimate_{uuid.uuid4().hex[:8]}.xlsx"
    filepath = os.path.join("/tmp", filename)
    wb.save(filepath)

    safe_name = req.customer_name.replace("/", "_").replace("\\", "_") if req.customer_name else "見積書"
    download_name = f"見積書_{safe_name}_{datetime.now().strftime('%Y%m%d')}.xlsx"

    return FileResponse(
        path=filepath,
        filename=download_name,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.post("/generate-spec", tags=["generate"])
async def generate_full_spec_excel(req: SpecSheetRequest):
    """
    仕様書Excelを生成してダウンロード（全フィールド版）

    フロントエンドのSpecSheetコンポーネントから全仕様フィールドを受け取り、
    セクション分けされた帳票Excelを生成します。
    """
    wb = _build_spec_workbook(req.model_dump())

    filename = f"spec_{uuid.uuid4().hex[:8]}.xlsx"
    filepath = os.path.join("/tmp", filename)
    wb.save(filepath)

    safe_name = req.project.replace("/", "_").replace("\\", "_") if req.project else "仕様書"
    download_name = f"仕様書_{safe_name}_{datetime.now().strftime('%Y%m%d')}.xlsx"

    return FileResponse(
        path=filepath,
        filename=download_name,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# ============================================================
# 見積書 PDF 生成
# ============================================================

# フォントパス検索
_FONT_PATH: Optional[str] = None
for _p in [
    "/usr/share/fonts/truetype/droid/DroidSansFallbackFull.ttf",
    "/usr/share/fonts/truetype/noto/NotoSansCJK-Regular.ttc",
    "/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc",
]:
    if Path(_p).exists():
        _FONT_PATH = _p
        break


class _EstimatePDF(FPDF):
    """見積書PDF生成クラス"""

    def __init__(self, font_path: Optional[str] = None):
        super().__init__(orientation="P", unit="mm", format="A4")
        self._has_jp = False
        if font_path and Path(font_path).exists():
            self.add_font("DroidJP", "", font_path)
            self.add_font("DroidJP", "B", font_path)
            self._has_jp = True
        self.set_auto_page_break(auto=False)

    def _jp(self, size: int = 9, bold: bool = False):
        if self._has_jp:
            self.set_font("DroidJP", "B" if bold else "", size)
        else:
            self.set_font("Helvetica", "B" if bold else "", size)

    def _ascii(self, size: int = 9, bold: bool = False):
        self.set_font("Helvetica", "B" if bold else "", size)

    @staticmethod
    def _fmt(val: Any) -> str:
        if val is None or val == 0 or val == "":
            return ""
        return f"{int(val):,}"


def _build_estimate_pdf(data: dict) -> str:
    """見積書PDFを生成してファイルパスを返す"""
    pdf = _EstimatePDF(font_path=_FONT_PATH)

    customer = data.get("customer_name", "")
    today_str = data.get("estimate_date", datetime.now().strftime("%Y-%m-%d"))
    expiry = ""
    try:
        dt = datetime.fromisoformat(today_str)
        expiry = (dt + timedelta(days=30)).strftime("%Y-%m-%d")
    except Exception:
        expiry = ""

    # オプション工事明細
    opt_labels = [
        "仮設工事", "基礎工事", "躯体工事", "屋根・板金工事", "外壁工事",
        "建具工事", "内装工事", "電気設備工事", "給排水衛生設備", "空調換気設備",
    ]
    opt_keys = [f"opt_{i:02d}_amount" for i in range(1, 11)]
    opt_items = [(lbl, data.get(k)) for lbl, k in zip(opt_labels, opt_keys)]

    # 付帯工事明細
    fut_labels = [
        "付帯工事", "諸経費", "エアコン工事", "防水工事",
        "太陽光", "蓄電池", "EV充電器", "外構工事",
    ]
    fut_keys = [f"futai_{i:02d}_amount" for i in range(1, 9)]
    fut_items = [(lbl, data.get(k)) for lbl, k in zip(fut_labels, fut_keys)]

    # 金額集計
    body_price = data.get("body_price_detail") or data.get("body_price") or 0
    opt_total = sum(v or 0 for _, v in opt_items)
    fut_total = sum(v or 0 for _, v in fut_items)
    contract = (body_price or 0) + opt_total + fut_total
    tax = int(contract * 0.1)
    total = contract + tax

    # 上書き値があればそちらを使用
    if data.get("contract_amount"):
        contract = data["contract_amount"]
    if data.get("tax"):
        tax = data["tax"]
    if data.get("total_with_tax"):
        total = data["total_with_tax"]
    if data.get("option_subtotal"):
        opt_total = data["option_subtotal"]
    if data.get("futai_subtotal"):
        fut_total = data["futai_subtotal"]

    # ===== ページ1: 見積書サマリー =====
    pdf.add_page()

    # ヘッダー情報（右上）
    pdf._jp(7)
    pdf.set_xy(120, 10); pdf.cell(20, 5, "見積番号", align="R")
    pdf._ascii(8)
    pdf.set_xy(140, 10); pdf.cell(50, 5, "No.0001", align="R")
    pdf._jp(7)
    pdf.set_xy(120, 15); pdf.cell(20, 5, "作成日", align="R")
    pdf._ascii(8)
    pdf.set_xy(140, 15); pdf.cell(50, 5, today_str, align="R")
    pdf._jp(7)
    pdf.set_xy(120, 20); pdf.cell(20, 5, "有効期限", align="R")
    pdf._ascii(8)
    pdf.set_xy(140, 20); pdf.cell(50, 5, expiry, align="R")

    # タイトル
    pdf._jp(18, bold=True)
    pdf.set_xy(15, 30); pdf.cell(170, 12, "御  見  積  書", align="C")

    # お客様名
    pdf._jp(14)
    pdf.set_xy(20, 48); pdf.cell(80, 10, customer, align="L")
    pdf._jp(10)
    pdf.cell(10, 10, "様", align="L")
    pdf.line(20, 58, 110, 58)

    # 見積金額ボックス
    pdf.set_fill_color(245, 245, 245)
    pdf.set_draw_color(80)
    pdf.set_line_width(0.4)
    pdf.rect(20, 64, 170, 18, "FD")
    pdf.set_line_width(0.2)
    pdf._jp(10)
    pdf.set_xy(25, 67); pdf.cell(40, 5, "建物基本本体価格", align="L")
    pdf._ascii(20, bold=True)
    pdf.set_xy(70, 65); pdf.cell(90, 16, pdf._fmt(total), align="R")
    pdf._jp(10)
    pdf.set_xy(162, 74); pdf.cell(20, 5, "円（税込）", align="L")

    # 内訳
    y = 90
    pdf._jp(10, bold=True)
    pdf.set_xy(25, y); pdf.cell(40, 7, "内訳", align="L")
    y += 10

    for label, val in [
        ("建物基本本体価格", body_price),
        ("オプション工事価格", opt_total),
        ("付帯工事価格", fut_total),
    ]:
        pdf._jp(9)
        pdf.set_xy(30, y); pdf.cell(60, 7, label, align="L")
        pdf._ascii(9)
        pdf.set_xy(100, y); pdf.cell(60, 7, pdf._fmt(val), align="R")
        pdf._jp(9)
        pdf.set_xy(162, y); pdf.cell(10, 7, "円", align="L")
        y += 8

    pdf.set_draw_color(0)
    pdf.line(30, y, 175, y)
    y += 4

    for label, val, bold in [
        ("小計", contract, False),
        ("消費税", tax, False),
    ]:
        pdf._jp(9, bold=bold)
        pdf.set_xy(30, y); pdf.cell(60, 7, label, align="L")
        pdf._ascii(9, bold=bold)
        pdf.set_xy(100, y); pdf.cell(60, 7, pdf._fmt(val), align="R")
        pdf._jp(9)
        pdf.set_xy(162, y); pdf.cell(10, 7, "円", align="L")
        y += 8

    # 二重線
    pdf.set_line_width(0.5)
    pdf.line(30, y, 175, y)
    pdf.set_line_width(0.2)
    pdf.line(30, y + 1, 175, y + 1)
    y += 5

    # 請負金額
    pdf._jp(11, bold=True)
    pdf.set_xy(30, y); pdf.cell(60, 8, "請負金額", align="L")
    pdf._ascii(11, bold=True)
    pdf.set_xy(100, y); pdf.cell(60, 8, pdf._fmt(total), align="R")
    pdf._jp(11)
    pdf.set_xy(162, y); pdf.cell(10, 8, "円", align="L")
    y += 15

    # 床面積テーブル
    f1s = data.get("floor_1f_sqm")
    if f1s is not None:
        pdf._jp(9, bold=True)
        pdf.set_xy(25, y); pdf.cell(60, 7, "物件概要 / 床面積", align="L")
        y += 9
        pdf.set_fill_color(220, 225, 235)
        for label, w in [("", 20), ("㎡", 30), ("坪", 30)]:
            pdf._jp(8)
            pdf.set_xy(30 + [("", 20), ("㎡", 30), ("坪", 30)].index((label, w)) * [20, 30, 30][[("", 20), ("㎡", 30), ("坪", 30)].index((label, w))], y)
        # シンプルにリライト
        x0 = 30
        for label, w in [("", 20), ("㎡", 30), ("坪", 30)]:
            pdf._jp(8)
            pdf.set_xy(x0, y)
            pdf.cell(w, 6, label, border=1, align="C", fill=True)
            x0 += w
        y += 6

        floor_rows = [
            ("1F", data.get("floor_1f_sqm"), data.get("floor_1f_tsubo")),
            ("2F", data.get("floor_2f_sqm"), data.get("floor_2f_tsubo")),
            ("合計", data.get("floor_total_sqm"), data.get("floor_total_tsubo")),
        ]
        for label, sqm, tsubo in floor_rows:
            x0 = 30
            pdf._jp(8)
            pdf.set_xy(x0, y); pdf.cell(20, 6, label, border=1, align="C")
            pdf._ascii(8)
            pdf.set_xy(x0 + 20, y); pdf.cell(30, 6, str(sqm) if sqm else "", border=1, align="R")
            pdf.set_xy(x0 + 50, y); pdf.cell(30, 6, str(tsubo) if tsubo else "", border=1, align="R")
            y += 6

    # ===== ページ2: 工事明細 =====
    pdf.add_page()
    y = 15

    def _detail_table(y_pos: int, title: str, items: list) -> int:
        pdf._jp(9, bold=True)
        pdf.set_xy(25, y_pos); pdf.cell(40, 7, title, align="L")
        y_pos += 8

        pdf.set_fill_color(200, 210, 230)
        pdf._jp(7)
        pdf.set_xy(25, y_pos); pdf.cell(65, 6, "工事内容", border=1, align="C", fill=True)
        pdf.set_xy(90, y_pos); pdf.cell(45, 6, "金額", border=1, align="C", fill=True)
        y_pos += 6

        subtotal = 0
        for name, amount in items:
            if amount is None or amount == 0:
                continue
            pdf._jp(8)
            pdf.set_xy(25, y_pos); pdf.cell(65, 7, name, border=1, align="L")
            pdf._ascii(8)
            pdf.set_xy(90, y_pos); pdf.cell(45, 7, pdf._fmt(amount), border=1, align="R")
            subtotal += amount
            y_pos += 7

        pdf.set_fill_color(240, 240, 240)
        pdf._jp(8, bold=True)
        pdf.set_xy(25, y_pos); pdf.cell(65, 7, f"{title}  小計", border=1, align="R", fill=True)
        pdf._ascii(8, bold=True)
        pdf.set_xy(90, y_pos); pdf.cell(45, 7, pdf._fmt(subtotal), border=1, align="R", fill=True)
        return y_pos + 12

    # オプション工事テーブル
    active_opts = [(n, a) for n, a in opt_items if a and a > 0]
    if active_opts:
        y = _detail_table(y, "オプション工事", active_opts)

    # 付帯工事テーブル
    active_futs = [(n, a) for n, a in fut_items if a and a > 0]
    if active_futs:
        y = _detail_table(y, "付帯工事", active_futs)

    # ファイル保存
    filename = f"estimate_{uuid.uuid4().hex[:8]}.pdf"
    filepath = os.path.join("/tmp", filename)
    pdf.output(filepath)
    return filepath


@app.post("/generate-estimate-pdf", tags=["generate"])
async def generate_estimate_pdf(req: EstimateRequest):
    """
    見積書PDFを生成してダウンロード

    レイアウト固定のPDFを直接生成します。
    お客様提出用として雛形が崩れません。
    """
    data = req.model_dump()
    filepath = _build_estimate_pdf(data)

    safe_name = req.customer_name.replace("/", "_").replace("\\", "_") if req.customer_name else "見積書"
    download_name = f"見積書_{safe_name}_{datetime.now().strftime('%Y%m%d')}.pdf"

    return FileResponse(
        path=filepath,
        filename=download_name,
        media_type="application/pdf",
    )


# ルータを組み込む
app.include_router(api_router, prefix=settings.API_V1_PREFIX)


if __name__ == "__main__":
    import uvicorn

    uvicorn.run(
        app,
        host="0.0.0.0",
        port=8000,
        log_level=settings.LOG_LEVEL.lower(),
        reload=settings.DEBUG,
    )
