"""
ExcelService テスト

Excelファイル生成ロジックのユニットテストです。
"""

import asyncio
import tempfile
from pathlib import Path

import pytest
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from app.models.template import CellMapping
from app.services.excel_service import ExcelService
from app.services.storage_service import LocalStorage


@pytest.fixture
def sample_template_path(tmp_path):
    """
    サンプルテンプレートを作成

    96行×20列のテンプレートを生成します。
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # テンプレートデータを埋め込む
    for row in range(1, 97):
        for col in range(1, 21):
            cell = ws.cell(row=row, column=col)
            cell.value = f"Cell_{row}_{col}"

    # マージセルを追加
    ws.merge_cells("A1:C3")
    ws["A1"] = "Merged Cell"

    # ファイルを保存
    template_file = tmp_path / "test_template.xlsx"
    wb.save(str(template_file))

    return str(template_file)


@pytest.fixture
def sample_cell_mappings():
    """
    サンプルセルマッピングを作成
    """
    return [
        CellMapping(
            id=1,
            template_id=1,
            cell_ref="A1",
            field_name="company_name",
            data_type="string",
            sort_order=1,
        ),
        CellMapping(
            id=2,
            template_id=1,
            cell_ref="B2",
            field_name="invoice_number",
            data_type="string",
            sort_order=2,
        ),
        CellMapping(
            id=3,
            template_id=1,
            cell_ref="C3",
            field_name="amount",
            data_type="float",
            sort_order=3,
        ),
    ]


@pytest.fixture
def sample_input_data():
    """
    サンプル入力データ
    """
    return {
        "company_name": "株式会社ABC",
        "invoice_number": "INV-001",
        "amount": 1000000.50,
    }


@pytest.fixture
def storage_service(tmp_path):
    """
    ストレージサービスインスタンス
    """
    return LocalStorage(base_path=str(tmp_path))


@pytest.fixture
def excel_service(storage_service):
    """
    ExcelService インスタンス
    """
    return ExcelService(storage_service)


class TestExcelService:
    """ExcelService テストクラス"""

    @pytest.mark.asyncio
    async def test_load_template(self, excel_service, sample_template_path):
        """
        テンプレート読み込みテスト

        テンプレートファイルが正しく読み込めることを確認します。
        """
        # テンプレートを読み込む
        workbook = await excel_service.load_template(sample_template_path)

        # 確認
        assert workbook is not None
        assert workbook.active.title == "Sheet1"

    @pytest.mark.asyncio
    async def test_load_template_not_found(self, excel_service):
        """
        テンプレート不在テスト

        存在しないテンプレートを読み込もうとした場合のエラーを確認します。
        """
        with pytest.raises(FileNotFoundError):
            await excel_service.load_template("nonexistent_template.xlsx")

    @pytest.mark.asyncio
    async def test_fill_cells(self, excel_service, sample_template_path, sample_cell_mappings, sample_input_data):
        """
        セルデータ埋め込みテスト

        セルマッピングに基づいてデータが正しく埋め込まれることを確認します。
        """
        # テンプレートを読み込む
        workbook = await excel_service.load_template(sample_template_path)
        ws = workbook.active

        # セルを埋め込む
        filled_wb = await excel_service.fill_cells(
            workbook,
            sample_cell_mappings,
            sample_input_data,
        )

        # 確認
        assert filled_wb["A1"].value == "株式会社ABC"
        assert filled_wb["B2"].value == "INV-001"
        assert filled_wb["C3"].value == 1000000.50

    @pytest.mark.asyncio
    async def test_fill_cells_with_merged_cells(self, excel_service, sample_template_path, sample_input_data):
        """
        マージセル埋め込みテスト

        マージされたセルにデータが正しく埋め込まれることを確認します。
        """
        # テンプレートを読み込む
        workbook = await excel_service.load_template(sample_template_path)
        ws = workbook.active

        # マージセルのマッピング
        mappings = [
            CellMapping(
                id=1,
                template_id=1,
                cell_ref="A1",  # マージ範囲の左上セル
                field_name="header_text",
                data_type="string",
                sort_order=1,
            ),
        ]

        input_data = {"header_text": "重要な情報"}

        # セルを埋め込む
        filled_wb = await excel_service.fill_cells(
            workbook,
            mappings,
            input_data,
        )

        # マージセル（A1）に値が埋め込まれていることを確認
        assert filled_wb["A1"].value == "重要な情報"

    @pytest.mark.asyncio
    async def test_safe_write_cell_finds_merge_range(self, excel_service, sample_template_path):
        """
        マージセル検出テスト

        マージセルの範囲が正しく検出されることを確認します。
        """
        # テンプレートを読み込む
        workbook = await excel_service.load_template(sample_template_path)
        ws = workbook.active

        # A1:C3 がマージされているので、B1 を指定してもA1に書き込まれるはず
        await excel_service._safe_write_cell(ws, "B1", "Test Value")

        # A1 の値が変わっていることを確認
        assert ws["A1"].value == "Test Value"

    @pytest.mark.asyncio
    async def test_find_merge_range_top_left(self, excel_service, sample_template_path):
        """
        マージ範囲左上セル検出テスト

        マージセルの左上セルが正しく検出されることを確認します。
        """
        workbook = await excel_service.load_template(sample_template_path)
        ws = workbook.active

        # A1:C3 がマージされているため、B2 を指定するとA1を返すはず
        top_left = excel_service._find_merge_range_top_left(ws, "B2")
        assert top_left == "A1"

        # マージされていないセルを指定すると None を返すはず
        top_left = excel_service._find_merge_range_top_left(ws, "D4")
        assert top_left is None

    @pytest.mark.asyncio
    async def test_data_type_conversion_int(self, excel_service, sample_template_path):
        """
        整数型変換テスト

        入力データが整数型に変換されることを確認します。
        """
        workbook = await excel_service.load_template(sample_template_path)

        mappings = [
            CellMapping(
                id=1,
                template_id=1,
                cell_ref="A1",
                field_name="quantity",
                data_type="int",
                sort_order=1,
            ),
        ]

        input_data = {"quantity": "100"}  # 文字列

        filled_wb = await excel_service.fill_cells(
            workbook,
            mappings,
            input_data,
        )

        # 整数型に変換されているはず
        assert filled_wb["A1"].value == 100
        assert isinstance(filled_wb["A1"].value, int)

    @pytest.mark.asyncio
    async def test_data_type_conversion_float(self, excel_service, sample_template_path):
        """
        浮動小数点型変換テスト

        入力データが浮動小数点型に変換されることを確認します。
        """
        workbook = await excel_service.load_template(sample_template_path)

        mappings = [
            CellMapping(
                id=1,
                template_id=1,
                cell_ref="A1",
                field_name="price",
                data_type="float",
                sort_order=1,
            ),
        ]

        input_data = {"price": "99.99"}  # 文字列

        filled_wb = await excel_service.fill_cells(
            workbook,
            mappings,
            input_data,
        )

        # 浮動小数点型に変換されているはず
        assert filled_wb["A1"].value == 99.99
        assert isinstance(filled_wb["A1"].value, float)


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
