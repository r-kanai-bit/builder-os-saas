"""Vercel Serverless Function: 見積書Excel生成（テンプレートベース）"""
from http.server import BaseHTTPRequestHandler
import json, os, io, base64
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook
from typing import Optional, Any

# テンプレートパス（api/ ディレクトリ内）
TEMPLATE_PATH = Path(__file__).parent / "estimate_v2.xlsx"

# セルマッピング（36フィールド）
CELL_MAPPINGS = {
    "customer_name": "B5", "estimate_date": "N5",
    "total_with_tax_2": "K10",
    "body_price": "N20", "option_total": "N22", "futai_total": "N24",
    "contract_amount": "N27", "tax": "N30", "total_with_tax": "N33",
    "body_price_detail": "X65",
    "opt_01_amount": "U69", "opt_02_amount": "U70", "opt_03_amount": "U71",
    "opt_04_amount": "U72", "opt_05_amount": "U73", "opt_06_amount": "U74",
    "opt_07_amount": "U75", "opt_08_amount": "U76", "opt_09_amount": "U77",
    "opt_10_amount": "U78", "option_subtotal": "N82",
    "futai_01_amount": "U86", "futai_02_amount": "U87", "futai_03_amount": "U88",
    "futai_04_amount": "U89", "futai_05_amount": "U90", "futai_06_amount": "U91",
    "futai_07_amount": "U92", "futai_08_amount": "U93", "futai_subtotal": "N96",
    "floor_1f_sqm": "E48", "floor_1f_tsubo": "I48",
    "floor_2f_sqm": "E49", "floor_2f_tsubo": "I49",
    "floor_total_sqm": "E51", "floor_total_tsubo": "I51",
}


def _find_merge_top_left(ws, cell_ref: str) -> Optional[str]:
    for merged_range in ws.merged_cells.ranges:
        if cell_ref in merged_range:
            return merged_range.start_cell.coordinate
    return None


def _safe_write(ws, cell_ref: str, value: Any) -> None:
    top_left = _find_merge_top_left(ws, cell_ref)
    ws[top_left or cell_ref] = value


class handler(BaseHTTPRequestHandler):
    def do_POST(self):
        content_length = int(self.headers.get("Content-Length", 0))
        body = json.loads(self.rfile.read(content_length)) if content_length else {}

        if not TEMPLATE_PATH.exists():
            self.send_response(500)
            self.send_header("Content-Type", "application/json")
            self.end_headers()
            self.wfile.write(json.dumps({"error": "テンプレートが見つかりません"}).encode())
            return

        wb = load_workbook(str(TEMPLATE_PATH), data_only=False)
        ws = wb.active

        for field_name, cell_ref in CELL_MAPPINGS.items():
            value = body.get(field_name)
            if value is not None:
                if field_name == "estimate_date" and isinstance(value, str) and value:
                    try:
                        value = datetime.fromisoformat(value)
                    except ValueError:
                        pass
                _safe_write(ws, cell_ref, value)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        data = buf.getvalue()

        safe_name = (body.get("customer_name") or "見積書").replace("/", "_").replace("\\", "_")
        filename = f"見積書_{safe_name}_{datetime.now().strftime('%Y%m%d')}.xlsx"

        self.send_response(200)
        self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        self.send_header("Content-Disposition", f'attachment; filename="{filename}"')
        self.send_header("Content-Length", str(len(data)))
        self.send_header("Access-Control-Allow-Origin", "*")
        self.end_headers()
        self.wfile.write(data)

    def do_OPTIONS(self):
        self.send_response(200)
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Access-Control-Allow-Methods", "POST, OPTIONS")
        self.send_header("Access-Control-Allow-Headers", "Content-Type")
        self.end_headers()
